import os
import json
import re
import tempfile
import anthropic
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY")

client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

SEARCH_PROMPT = """You are a fitness research assistant. Search online for scientific workout programs and evidence-based exercises. Find the best exercises for the user's goals. Return a brief summary in plain text of what you found."""

JSON_PROMPT = """You are an expert fitness coach. Based on the research provided, create a personalized workout plan. Reply ONLY with a valid JSON object. Rules:
- Use ONLY ASCII characters (a-z, A-Z, 0-9, spaces, basic punctuation)
- NO accents, NO special characters, NO unicode
- Keep each text field under 10 words
- Follow EXACTLY what the user requested: days, muscle groups, level, injuries
- No text before or after the JSON

JSON structure:
{"nome_utente":"","obiettivo":"","livello":"","giorni_settimana":4,"metodologia":"","note_generali":"","progressione":"","giorni":[{"giorno":"Monday","focus":"Chest","esercizi":[{"nome":"Bench Press","serie":4,"ripetizioni":"6-8","recupero":"2min","note":"control descent arch back","perche":"best chest compound"}]}]}"""


def pulisci_json(testo):
    testo = testo.replace("```json","").replace("```","").strip()
    testo = re.sub(r'[\x00-\x1f\x7f-\xff]', lambda m: '' if ord(m.group()) > 127 else m.group(), testo)
    testo = re.sub(r',\s*}', '}', testo)
    testo = re.sub(r',\s*]', ']', testo)
    match = re.search(r'\{.*\}', testo, re.DOTALL)
    if not match:
        raise ValueError("JSON not found")
    return json.loads(match.group())


def crea_excel(dati, nome_file):
    wb = Workbook()
    NERO="1A1A2E"; ROSSO="E63946"; GRIGIO_SC="2D2D2D"
    GRIGIO_CH="F5F5F5"; BIANCO="FFFFFF"; GIALLO="FFD700"
    COLORI_GIORNI=["16213E","0F3460","533483","2C2C54","1B4332","3D0C02"]

    def fill(h): return PatternFill("solid",fgColor=h)
    def fnt(bold=False,color=BIANCO,size=11):
        return Font(bold=bold,color=color,name="Calibri",size=size)
    def brd():
        s=Side(style="thin",color="CCCCCC")
        return Border(left=s,right=s,top=s,bottom=s)
    def ctr(wrap=False):
        return Alignment(horizontal="center",vertical="center",wrap_text=wrap)
    def lft(wrap=False):
        return Alignment(horizontal="left",vertical="center",wrap_text=wrap)

    ws1=wb.active; ws1.title="Scheda Settimanale"
    ws1.sheet_view.showGridLines=False

    ws1.merge_cells("A1:H1"); ws1.row_dimensions[1].height=50
    c=ws1["A1"]
    c.value="FITCOACH PRO | Il tuo Personal Trainer AI"
    c.fill=fill(NERO)
    c.font=Font(bold=True,color=GIALLO,name="Calibri",size=16)
    c.alignment=ctr()

    ws1.merge_cells("A2:H2"); ws1.row_dimensions[2].height=28
    c=ws1["A2"]
    c.value=(str(dati.get("nome_utente",""))+" | "+
             str(dati.get("obiettivo",""))+" | "+
             str(dati.get("livello",""))+" | "+
             str(dati.get("giorni_settimana",""))+" giorni/sett")
    c.fill=fill(ROSSO); c.font=fnt(bold=True,size=11); c.alignment=ctr()

    ws1.merge_cells("A3:H3"); ws1.row_dimensions[3].height=36
    c=ws1["A3"]
    c.value="METODOLOGIA: "+str(dati.get("metodologia",""))
    c.fill=fill(GRIGIO_SC); c.font=fnt(size=10); c.alignment=lft(wrap=True)

    headers=["GIORNO","FOCUS","ESERCIZIO","SERIE","REPS","RECUPERO","NOTE TECNICHE","PERCHE"]
    widths=[13,18,26,7,10,10,40,40]
    ws1.row_dimensions[4].height=30
    for col,(h,w) in enumerate(zip(headers,widths),1):
        ws1.column_dimensions[get_column_letter(col)].width=w
        c=ws1.cell(row=4,column=col,value=h)
        c.fill=fill(GRIGIO_SC); c.font=fnt(bold=True,size=10)
        c.alignment=ctr(); c.border=brd()

    row=5
    for idx,giorno in enumerate(dati.get("giorni",[])):
        col_g=COLORI_GIORNI[idx%len(COLORI_GIORNI)]
        esercizi=giorno.get("esercizi",[]); prima=row
        for i,ex in enumerate(esercizi):
            ws1.row_dimensions[row].height=42
            if i==0:
                if len(esercizi)>1:
                    ws1.merge_cells("A"+str(prima)+":A"+str(prima+len(esercizi)-1))
                    ws1.merge_cells("B"+str(prima)+":B"+str(prima+len(esercizi)-1))
                c=ws1.cell(row=prima,column=1,value=giorno.get("giorno",""))
                c.fill=fill(col_g); c.font=fnt(bold=True,size=10)
                c.alignment=ctr(wrap=True); c.border=brd()
                c=ws1.cell(row=prima,column=2,value=giorno.get("focus",""))
                c.fill=fill(col_g); c.font=fnt(bold=True,size=10)
                c.alignment=ctr(wrap=True); c.border=brd()
            bg=GRIGIO_CH if i%2==0 else BIANCO
            valori=[
                ex.get("nome",""),ex.get("serie",""),
                ex.get("ripetizioni",""),ex.get("recupero",""),
                ex.get("note",""),ex.get("perche","")
            ]
            for col_off,val in enumerate(valori):
                c=ws1.cell(row=row,column=3+col_off,value=val)
                c.fill=fill(bg)
                c.font=Font(color=GRIGIO_SC,name="Calibri",size=10,bold=(col_off==0))
                c.alignment=ctr() if col_off<4 else lft(wrap=True)
                c.border=brd()
            row+=1
        for col in range(1,9):
            c=ws1.cell(row=row,column=col,value="")
            c.fill=fill(NERO)
        ws1.row_dimensions[row].height=5; row+=1

    ws2=wb.create_sheet("Spiegazioni")
    ws2.sheet_view.showGridLines=False

    ws2.merge_cells("A1:D1"); ws2.row_dimensions[1].height=50
    c=ws2["A1"]
    c.value="FITCOACH PRO | Guida agli Esercizi"
    c.fill=fill(NERO)
    c.font=Font(bold=True,color=GIALLO,name="Calibri",size=15)
    c.alignment=ctr()

    hdrs2=["ESERCIZIO","GIORNO","NOTE TECNICHE","PERCHE QUESTO ESERCIZIO"]
    widths2=[26,22,50,50]
    ws2.row_dimensions[2].height=30
    for col,(h,w) in enumerate(zip(hdrs2,widths2),1):
        ws2.column_dimensions[get_column_letter(col)].width=w
        c=ws2.cell(row=2,column=col,value=h)
        c.fill=fill(GRIGIO_SC); c.font=fnt(bold=True,size=10)
        c.alignment=ctr(); c.border=brd()

    r2=3
    for idx,giorno in enumerate(dati.get("giorni",[])):
        col_g=COLORI_GIORNI[idx%len(COLORI_GIORNI)]
        for i,ex in enumerate(giorno.get("esercizi",[])):
            ws2.row_dimensions[r2].height=52
            bg=GRIGIO_CH if i%2==0 else BIANCO
            vals=[
                ex.get("nome",""),
                str(giorno.get("giorno",""))+" - "+str(giorno.get("focus","")),
                ex.get("note",""),
                ex.get("perche","")
            ]
            for col_idx,val in enumerate(vals,1):
                c=ws2.cell(row=r2,column=col_idx,value=val)
                c.fill=fill(col_g if col_idx==2 else bg)
                c.font=Font(
                    color=BIANCO if col_idx==2 else GRIGIO_SC,
                    name="Calibri",size=10,bold=(col_idx==1))
                c.alignment=ctr(wrap=True) if col_idx==2 else lft(wrap=True)
                c.border=brd()
            r2+=1

    ws3=wb.create_sheet("Progressione")
    ws3.sheet_view.showGridLines=False
    ws3.column_dimensions["A"].width=28
    ws3.column_dimensions["B"].width=72

    ws3.merge_cells("A1:B1"); ws3.row_dimensions[1].height=50
    c=ws3["A1"]
    c.value="FITCOACH PRO | Piano di Progressione"
    c.fill=fill(NERO)
    c.font=Font(bold=True,color=GIALLO,name="Calibri",size=16)
    c.alignment=ctr()

    sezioni=[
        ("Obiettivo",dati.get("obiettivo","")),
        ("Livello",dati.get("livello","")),
        ("Metodologia",dati.get("metodologia","")),
        ("Progressione 1-8 sett",dati.get("progressione","")),
        ("Nutrizione e Recupero",dati.get("note_generali",""))
    ]
    r3=2
    for titolo,valore in sezioni:
        ws3.row_dimensions[r3].height=max(32,min(120,len(str(valore))//2))
        c=ws3.cell(row=r3,column=1,value=titolo)
        c.fill=fill(ROSSO); c.font=fnt(bold=True,size=11)
        c.alignment=lft(); c.border=brd()
        c=ws3.cell(row=r3,column=2,value=valore)
        c.fill=fill(GRIGIO_CH)
        c.font=Font(color=GRIGIO_SC,name="Calibri",size=10)
        c.alignment=lft(wrap=True); c.border=brd()
        r3+=1

    footer="Creato da FitCoach Pro | Il tuo Personal Trainer AI su Telegram"
    for ws in [ws1,ws2,ws3]:
        last=ws.max_row+2
        ws.merge_cells("A"+str(last)+":H"+str(last))
        c=ws.cell(row=last,column=1,value=footer)
        c.fill=fill(NERO)
        c.font=Font(color=GIALLO,name="Calibri",size=9,italic=True)
        c.alignment=ctr()

    wb.save(nome_file)
    return nome_file


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Benvenuto in FitCoach Pro!\n\n"
        "Sono il tuo personal trainer AI.\n\n"
        "Dimmi:\n"
        "- Obiettivi (massa, forza, dimagrimento)\n"
        "- Punti deboli (es. tricipiti scarsi)\n"
        "- Giorni a settimana\n"
        "- Esperienza (principiante/intermedio/avanzato)\n"
        "- Infortuni o limitazioni\n\n"
        "Riceverai un file Excel con la tua scheda completa!"
    )


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Descrivimi la tua situazione!\n\n"
        "Esempio: Ho i tricipiti carenti, mi alleno 4 giorni, "
        "livello intermedio, voglio massa.\n\n"
        "Riceverai un Excel con la tua scheda!"
    )


async def crea_scheda(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_message = update.message.text
    user_name = update.message.from_user.first_name
    attesa = await update.message.reply_text(
        "Sto cercando le migliori schede e creando il tuo programma... circa 40 secondi!"
    )

    try:
        # PASSO 1: ricerca web
        search_response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1000,
            system=SEARCH_PROMPT,
            tools=[{"type": "web_search_20250305", "name": "web_search"}],
            messages=[{
                "role": "user",
                "content": "Search for the best scientific exercises and workout programs for: "+user_message
            }]
        )

        ricerca = ""
        for b in search_response.content:
            if hasattr(b, "text"):
                ricerca += b.text

        # PASSO 2: genera JSON pulito senza web search
        json_response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1500,
            system=JSON_PROMPT,
            messages=[{
                "role": "user",
                "content": (
                    "User name: "+user_name+"\n"
                    "User request: "+user_message+"\n"
                    "Research findings: "+ricerca[:500]+"\n\n"
                    "Now create the workout plan JSON. Follow EXACTLY the user request. ASCII only."
                )
            }]
        )

        testo = ""
        for b in json_response.content:
            if hasattr(b, "text"):
                testo += b.text

        dati = pulisci_json(testo)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            nome_file = tmp.name
        crea_excel(dati, nome_file)
        await attesa.delete()

        with open(nome_file, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename="FitCoachPro_"+user_name+".xlsx",
                caption=(
                    "Scheda pronta "+user_name+"!\n\n"
                    "3 fogli:\n"
                    "- Scheda Settimanale\n"
                    "- Spiegazioni\n"
                    "- Progressione"
                )
            )
        os.unlink(nome_file)

    except Exception as e:
        await attesa.delete()
        await update.message.reply_text(
            "Si e verificato un errore. Riprova o scrivi /start"
        )
        print("Errore: "+str(e))


def main():
    app = Application.builder().token(TELEGRAM_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, crea_scheda))
    print("Bot avviato!")
    app.run_polling()


if __name__ == "__main__":
    main()
